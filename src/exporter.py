"""
src/exporter.py

Saves the calculated DataFrames as the actual deliverable files the
problem statement requires:

    1. Master Performance File  -> data/processed/master_performance.xlsx
    2. Module Summary File      -> data/processed/module_summary.xlsx
    3. Final Rankings File      -> data/processed/final_rankings.xlsx

Each is also saved as .csv (lighter, easier to diff/version-control)
alongside the .xlsx (easier for non-technical people, e.g. the trainer,
to open and skim in Excel).

Excel files get light formatting (bold headers, autofit columns) so
they're presentable without extra manual work -- this is what turns
"a script that works" into "a deliverable someone can actually use."
"""

import pandas as pd
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def _autofit_and_style(filepath: str, sheet_name: str = 'Sheet1'):
    """
    Opens a just-saved .xlsx and applies basic formatting:
    bold white-on-blue header row, auto-sized columns.
    Pure cosmetics, but it's the difference between a deliverable
    and a debugging artifact.
    """
    import openpyxl
    wb = openpyxl.load_workbook(filepath)
    ws = wb[sheet_name]

    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    ws.freeze_panes = 'A2'  # keep header visible while scrolling
    wb.save(filepath)


def export_dataframe(df: pd.DataFrame, name: str, output_dir: str = 'data/processed/'):
    """
    Saves one DataFrame as both .csv and formatted .xlsx under output_dir,
    using `name` as the base filename (no extension).
    """
    out_path = Path(output_dir)
    out_path.mkdir(parents=True, exist_ok=True)

    csv_path = out_path / f'{name}.csv'
    xlsx_path = out_path / f'{name}.xlsx'

    df.to_csv(csv_path, index=False)
    df.to_excel(xlsx_path, index=False, sheet_name='Sheet1')
    _autofit_and_style(str(xlsx_path))

    print(f"  Saved: {csv_path}")
    print(f"  Saved: {xlsx_path}")

    return {'csv': str(csv_path), 'xlsx': str(xlsx_path)}


def export_all_outputs(results: dict, output_dir: str = 'data/processed/') -> dict:
    """
    Takes the dict returned by calculator.build_master_performance()
    and writes out all 3 required deliverable files.

    Master Performance File -> all students, all computed metrics
    Module Summary File     -> per-student, per-module breakdown
    Final Rankings File     -> master sheet sorted by rank, key columns only
    """
    paths = {}

    print("Exporting Master Performance File...")
    master_cols = [
        'rank', 'name', 'email', 'quizzes_attempted',
        'total_marks_scored', 'total_marks_possible',
        'avg_percentage', 'final_percentile', 'grade'
    ]
    paths['master'] = export_dataframe(
        results['master'][master_cols], 'master_performance', output_dir
    )

    print("\nExporting Module Summary File...")
    module_cols = [
        'name', 'email', 'module', 'marks_scored', 'marks_possible',
        'module_percentage', 'module_percentile'
    ]
    module_sorted = results['module'].sort_values(['module', 'module_percentage'], ascending=[True, False])
    paths['module'] = export_dataframe(
        module_sorted[module_cols], 'module_summary', output_dir
    )

    print("\nExporting Final Rankings File...")
    rankings_cols = ['rank', 'name', 'email', 'avg_percentage', 'grade']
    paths['rankings'] = export_dataframe(
        results['master'][rankings_cols], 'final_rankings', output_dir
    )

    return paths


if __name__ == '__main__':
    import sys
    sys.path.insert(0, '.')
    from src.quiz_parser import parse_all_quiz_files
    from src.roster import apply_roster
    from src.calculator import build_master_performance

    long_df = parse_all_quiz_files('data/raw/')
    long_df = apply_roster(long_df)
    results = build_master_performance(long_df)

    print("\n--- Exporting output files ---")
    paths = export_all_outputs(results)

    print(f"\nDone. {len(results['master'])} students processed.")
